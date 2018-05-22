import openpyxl
import re
import sys

from openpyxl.utils import get_column_letter

file = sys.argv[1]

excel_file = openpyxl.load_workbook(file)

evaluation_sheet = excel_file['evaluation']
summary_sheet = excel_file.create_sheet(title='throughput', index=1)
average_pattern = re.compile(".*average.*")
std_pattern = re.compile(".*StdD.*")
workload_index_pattern = re.compile("[1-5].*")
operation_pattern = re.compile("insert.*|read.*|scan.*")


def get_average_of_three_cells_in_column(sheet, column_index, row_start_index):
    column_name = get_column_letter(column_index)
    next_column_name = get_column_letter(column_index + 1)

    tmp = sheet[str(column_name) + str(row_start_index)].value
    if isinstance(tmp, str):
        if re.match(pattern=workload_index_pattern, string=tmp):
            tmp = tmp[2:]
        return {"avg": tmp}

    cell_sum = 0
    next_cell_sum = 0
    cnt = 0

    def is_number(value):
        return isinstance(value, int) or isinstance(value, float)

    for i in range(0, 3):
        cell_value = sheet[str(column_name) + str(row_start_index + i)].value
        next_cell_value = sheet[str(next_column_name) + str(row_start_index + i)].value

        if is_number(cell_value):
            cell_sum += cell_value
            cnt += 1

        if is_number(next_cell_value):
            next_cell_sum += next_cell_value

    if cnt != 0:
        avg = cell_sum / cnt
        if avg != 0 and re.match(pattern=operation_pattern, string=sheet[column_name + "1"].value):
            avg_sec = avg / 1000000
            avg_ops_per_sec = 1 / avg_sec

            if re.match(pattern=std_pattern, string=sheet[next_column_name + "1"].value):
                next_avg = next_cell_sum / cnt
                standard_derivation_percentage = next_avg / avg
                standard_derivation = avg_ops_per_sec * standard_derivation_percentage

                return {"avg": avg_ops_per_sec, "stdd": standard_derivation_percentage}

            return {"avg": avg_ops_per_sec}
        else:
            return {"avg": avg}
    else:
        return {"avg": tmp}


def should_take_column(letter_index, column_name):
    return letter_index <= 16 or re.match(pattern=average_pattern, string=column_name) and re.match(
        pattern=operation_pattern, string=column_name)


column_index = 1

for letter_index in range(1, evaluation_sheet.max_column + 1):
    letter = get_column_letter(letter_index)
    value = get_average_of_three_cells_in_column(evaluation_sheet, letter_index, 1)
    column_name = evaluation_sheet[letter + "1"].value

    if should_take_column(letter_index=letter_index, column_name=column_name):
        if letter_index <= 16:
            summary_sheet.cell(row=1, column=column_index, value=value["avg"])
            column_index += 1
        else:
            summary_sheet.cell(row=1, column=column_index, value="throughput " + value["avg"][:-4] + "(ops/sec)")
            summary_sheet.cell(row=1, column=column_index + 1, value="stdD " + value["avg"][:-4] + "%")
            column_index += 2

column_index = 1

for letter_index in range(1, evaluation_sheet.max_column + 1):
    letter = get_column_letter(letter_index)
    summary_row_index = 2
    column_name = evaluation_sheet[letter + "1"].value

    if should_take_column(letter_index=letter_index, column_name=column_name):
        to_advance = 1

        for row_index in range(5, 720, 6):
            average = get_average_of_three_cells_in_column(evaluation_sheet, letter_index, row_index)
            summary_sheet.cell(row=summary_row_index, column=column_index, value=average["avg"])

            if "stdd" in average:
                summary_sheet.cell(row=summary_row_index, column=column_index, value=average["avg"])
                summary_sheet.cell(row=summary_row_index, column=column_index + 1, value=average["stdd"])
                to_advance = 2
            else:
                summary_sheet.cell(row=summary_row_index, column=column_index, value=average["avg"])

            summary_row_index += 1

        column_index += to_advance

excel_file.save(filename=file)
